from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from nf_manager.models import Nf

def gerar_planilha_notas_emitidas(mes, ano):
    notas = Nf.objects.filter(
        data_emissao_nfse__year=ano,
        data_emissao_nfse__month=mes,
        status='emitido',
    )

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Notas Emitidas"

    headers = ["Data/Hora", "Empresa", "OP", "Valor", "Nota Fiscal"]

    for col, header in enumerate(headers, start=2):
        cell = sheet.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")


    for row, nota in enumerate(notas, start=7):
        sheet.cell(row=row, column=2, value=nota.data_emissao_nfse.replace(tzinfo=None))
        sheet.cell(row=row, column=3, value=nota.empresa.razao_social)
        sheet.cell(row=row, column=4, value=nota.op)
        sheet.cell(row=row, column=5, value=nota.valor)
        sheet.cell(row=row, column=6, value=nota.numero_nfse)

        sheet[f'E{row}'].number_format = 'R$ #,##0.00'

    sheet.freeze_panes = "B7"

    for column in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

    return wb